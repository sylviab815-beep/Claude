#!/usr/bin/env python3
"""Build a sortable, color-coded Excel sheet of genre communities.

Usage:
    python build_group_sheet.py results.json "Genre Name" output.xlsx

results.json: an array of objects with fields:
    platform, name, url, members, privacy, promo_allowed, focus, activity, notes
Missing values may be null or "unknown".
"""
import json
import sys


def load(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "groups" in data:
        data = data["groups"]
    return data


def members_key(g):
    m = g.get("members")
    return m if isinstance(m, (int, float)) else -1


def promo_color(value):
    """Return a fill hex based on whether promo is allowed."""
    if not value or value == "unknown" or str(value).lower().startswith("unknown"):
        return "FFF2CC"  # amber - verify
    v = str(value).lower()
    if v.startswith("no") or "ban" in v or "not allow" in v:
        return "F4CCCC"  # red - engagement only
    return "D9EAD3"  # green - promo allowed


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    results_path, genre, out_path = sys.argv[1], sys.argv[2], sys.argv[3]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.stderr.write("openpyxl not installed. Run: pip install openpyxl --break-system-packages\n")
        sys.exit(2)

    groups = load(results_path)
    # Sort: platform asc, then members desc
    groups.sort(key=lambda g: (str(g.get("platform", "")), -members_key(g)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Groups"

    headers = ["Platform", "Group Name", "URL", "Members", "Privacy",
               "Promo Allowed", "Focus", "Activity", "Notes"]
    fields = ["platform", "name", "url", "members", "privacy",
              "promo_allowed", "focus", "activity", "notes"]

    title_fill = PatternFill("solid", fgColor="38761D")
    header_fill = PatternFill("solid", fgColor="B6D7A8")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tcell = ws.cell(row=1, column=1, value=f"{genre} — Facebook & Goodreads Groups")
    tcell.font = Font(bold=True, size=14, color="FFFFFF")
    tcell.fill = title_fill
    tcell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows
    r = 3
    for g in groups:
        for c, field in enumerate(fields, start=1):
            val = g.get(field)
            if val is None:
                val = ""
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(field == "notes"))
            if field == "url" and val:
                cell.hyperlink = val
                cell.font = Font(color="1155CC", underline="single")
            if field == "members" and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            if field == "promo_allowed":
                cell.fill = PatternFill("solid", fgColor=promo_color(val))
        r += 1

    widths = [12, 34, 40, 11, 10, 26, 10, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(r-1,2)}"

    # Legend
    legend_row = r + 1
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    legend = [("Promo allowed", "D9EAD3"), ("Verify rules", "FFF2CC"), ("Engagement only / no promo", "F4CCCC")]
    for i, (label, color) in enumerate(legend):
        cell = ws.cell(row=legend_row + 1 + i, column=1, value=label)
        cell.fill = PatternFill("solid", fgColor=color)

    wb.save(out_path)
    print(f"Wrote {len(groups)} groups to {out_path}")


if __name__ == "__main__":
    main()
